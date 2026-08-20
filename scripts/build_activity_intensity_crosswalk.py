"""
Build the COICOP activity-intensity crosswalk and enrich Table A2.

Run AFTER extract_clean_dataset_from_workbook.py, which must have already
produced data/baseline_categories.csv and supplementary/table_a2_scenario4_parameters.csv.

Inputs
------
- source/Household_consumption_expenditure_ESA2010_current_prices_SEK_million_by_purpose_COICOP_and_year.xlsx
  Statistics Sweden, Statistikdatabasen table 000000SG, "Household consumption
  expenditure (ESA2010) by purpose COICOP (1999), 1980-2021." Download from
  https://www.statistikdatabasen.scb.se/pxweb/en/ssd/START__NR__NR0103__NR0103E/NR0103ENS2010T03NA/
  and place in source/. The table was frozen after May 2024 when SCB switched
  to COICOP 2018, so 2021 is its final year, not a stale download.
- data/baseline_categories.csv (produced by extract_clean_dataset_from_workbook.py)
- supplementary/table_a2_scenario4_parameters.csv (produced by the same script)

Outputs
-------
- data/coicop_activity_intensity_crosswalk.csv (all 109 active categories)
- supplementary/table_a2_scenario4_parameters.csv (overwritten, six columns added)

Every category-to-COICOP match and every "SHARED"/"LOW-CONF"/"DIRECT"/
"PHYSICAL"/"NOMATCH" judgement call below was made by hand against the COICOP
classification and against Table 1 and Tables S1-S5 of Dawkins et al. (2024)
and its supplementary material. See manuscript Section 3.8 for the narrative
account and the two documented limitations (national-expenditure proxy,
2021-vs-2022 vintage gap) that this file does not resolve, only makes visible.
"""
from pathlib import Path
import csv
import openpyxl

REPO_ROOT = Path(__file__).resolve().parents[1]
SOURCE_DIR = REPO_ROOT / 'source'
DATA_DIR = REPO_ROOT / 'data'
SUPP_DIR = REPO_ROOT / 'supplementary'

SCB_WORKBOOK = SOURCE_DIR / 'Household_consumption_expenditure_ESA2010_current_prices_SEK_million_by_purpose_COICOP_and_year.xlsx'
POP_2021 = 10_452_326  # Statistics Sweden, population 31 Dec 2021

# ---------------------------------------------------------------------------
# 1. Read national COICOP expenditure (SEK million, 2021) from the SCB workbook
# ---------------------------------------------------------------------------

def read_scb_workbook(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    scb = {}
    # SCB Statistikdatabasen exports use a two-column layout: label, value.
    # 5-digit codes appear at the start of the label for the finest rows we
    # use (e.g. "02131 Strong beer"); 3-4 digit codes elsewhere. We match on
    # the codes we need explicitly rather than parsing every row, since the
    # export includes header/footer metadata rows with no code.
    import re
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True):
        label, value = row[0], row[1]
        if not label or value is None:
            continue
        m = re.match(r'^(\d{3,5})\s+(.*)$', str(label).strip())
        if m:
            code, name = m.group(1), m.group(2)
            scb[code] = (name, float(value))
    return scb

# ---------------------------------------------------------------------------
# 2. Manuscript category -> COICOP code(s) crosswalk
#    flag: OK (clean match) | LOW-CONF (approximate proxy, see note) |
#          SHARED (COICOP reports several of our categories under one code;
#                  split proportionally to baseline emissions share) |
#          DIRECT (physical combustion, never expenditure-based) |
#          PHYSICAL (flights: physical km, not SEK) | NOMATCH (no COICOP line)
# ---------------------------------------------------------------------------

M = {}
def add(cat, codes=None, flag="OK", note="", peers=None):
    M[cat] = dict(codes=codes or [], flag=flag, note=note, peers=peers or [])

# Food and drink
add("Food and drink - Meat", ["0112"])
add("Food and drink - Milk, cheese, other dairy products and eggs", ["0114"])
add("Food and drink - Vegetables, root crops and legumes", ["0117"])
add("Food and drink - Fruit and nuts", ["0116"])
add("Food and drink - Bread and cereal products", ["0111"])
add("Food and drink - Sugar, confectionery and desserts", ["0118"])
add("Food and drink - Ready-made meals and other food products", ["0119"], flag="LOW-CONF",
    note="Best available proxy (salt/spices/sauces/baby food); ready-meals not separately coded at 4-digit COICOP")
add("Food and drink - Fish and seafood", ["0113"])
add("Food and drink - Oils and fats", ["0115"])
add("Food and drink - Wine", ["0212"])
add("Food and drink - Strong beer", ["02131"])
add("Food and drink - Low- and medium-strength beer", ["02132"])
add("Food and drink - Spirits", ["0211"])
add("Food and drink - Tobacco", ["022"])
add("Food and drink - Narcotics", ["023"])
add("Food and drink - Water", ["0122"], flag="SHARED", note="0122 combined with Soft drinks & Fruit/veg juices",
    peers=["Food and drink - Soft drinks", "Food and drink - Fruit and vegetable juices"])
add("Food and drink - Soft drinks", ["0122"], flag="SHARED", note="0122 combined with Water & Fruit/veg juices",
    peers=["Food and drink - Water", "Food and drink - Fruit and vegetable juices"])
add("Food and drink - Fruit and vegetable juices", ["0122"], flag="SHARED", note="0122 combined with Water & Soft drinks",
    peers=["Food and drink - Water", "Food and drink - Soft drinks"])
add("Food and drink - Coffee and coffee substitutes", ["0121"], flag="SHARED", note="0121 combined with Tea & Chocolate drinks",
    peers=["Food and drink - Tea, mat\u00e9 and other herbal teas", "Food and drink - Chocolate drinks"])
add("Food and drink - Tea, mat\u00e9 and other herbal teas", ["0121"], flag="SHARED", note="0121 combined with Coffee & Chocolate drinks",
    peers=["Food and drink - Coffee and coffee substitutes", "Food and drink - Chocolate drinks"])
add("Food and drink - Chocolate drinks", ["0121"], flag="SHARED", note="0121 combined with Coffee & Tea",
    peers=["Food and drink - Coffee and coffee substitutes", "Food and drink - Tea, mat\u00e9 and other herbal teas"])
add("Food and drink - Other alcoholic beverages", [], flag="NOMATCH", note="021 fully accounted by spirits+wine+beer; no residual COICOP line")

# Clothing and shoes
add("Clothing and shoes - Clothing", ["0312"])
add("Clothing and shoes - Shoes and other footwear", ["0321"])
add("Clothing and shoes - Other garments and clothing accessories", ["0313"])
add("Clothing and shoes - Clothing materials", ["0311"])
add("Clothing and shoes - Laundry, repair and rental of clothing", ["0314"])
add("Clothing and shoes - Cleaning, repair and rental of footwear", ["0322"])

# Housing
add("Housing - Actual rent for rented permanent dwelling", ["0411"])
add("Housing - Actual rents for other dwellings, garages and storage", ["0412"], flag="LOW-CONF",
    note="0412 mixes secondary-residence rentals paid & tenant-owner payments; approximate")
add("Housing - Imputed rent for owned permanent dwelling", ["0421"])
add("Housing - Imputed rent for holiday homes, other dwellings, garages and storage", ["0422"])
add("Housing - Maintenance, repair and security", ["043"])
add("Housing - Electricity", ["0451"])
add("Housing - Gas", ["0452"])
add("Housing - Liquid fuels", ["0453"])
add("Housing - Solid fuels", ["0454"])
add("Housing - District heating and other energy for heating and cooling", ["0455"], flag="LOW-CONF",
    note="0455 covers purchased district heat only; 'cooling' component not separately coded")
add("Housing - Direct emissions from heating", [], flag="DIRECT",
    note="Physical direct combustion (Statistics Sweden direct-emissions series), not COICOP expenditure. Use fuel volume, not SEK.")
add("Housing - Direct household emissions", [], flag="DIRECT",
    note="Physical direct combustion, not COICOP expenditure. Use fuel volume, not SEK.")

# Other consumption
add("Other consumption - Hotels, caf\u00e9s and restaurants", ["111"])
add("Other consumption - Accommodation services", ["112"])
add("Other consumption - Furniture, furnishings and carpets", ["0511", "0512"])
add("Other consumption - Information and communication equipment", ["0812", "0913"], flag="LOW-CONF",
    note="Combines telecom hardware (0812) + IT hardware (0913); approximate aggregate")
add("Other consumption - Information and communication services", ["0813"])
add("Other consumption - Medicines", ["0611"])
add("Other consumption - Medical products", ["0612"])
add("Other consumption - Medical aids and maintenance", ["0613"])
add("Other consumption - Other appliances and products for personal care", ["1213"])
add("Other consumption - Electric appliances for personal care", ["1212"])
add("Other consumption - Household appliances", ["0531", "0532"])
add("Other consumption - Goods and services for household maintenance", ["0561", "0562"])
add("Other consumption - Glassware, tableware and other utensils", ["054"])
add("Other consumption - Home and garden tools", ["0551", "0552"])
add("Other consumption - Household textiles", ["052"])
add("Other consumption - Financial services", ["126"])
add("Other consumption - Insurance", ["125"])
add("Other consumption - Hairdressing salons and personal grooming establishments", ["1211"])
add("Other consumption - Jewellery and watches", ["1231"])
add("Other consumption - Other personal effects n.e.c.", ["1232"])
add("Other consumption - Other services", ["127"], flag="LOW-CONF", note="Catch-all: legal/employment agency fees; imprecise proxy")
add("Other consumption - Outpatient dental services", ["0622"])
add("Other consumption - Other outpatient care services", ["0621", "0623"])
add("Other consumption - Inpatient curative and rehabilitative services", ["063"])
add("Other consumption - Home care for older people and people with disabilities", ["12402"], flag="LOW-CONF",
    note="12402 'care and help for elderly' only; disability home care not separately coded")
add("Other consumption - Residential care for older people and people with disabilities, excluding health care",
    ["12403"], flag="LOW-CONF", note="Best available proxy (personal assistance); residential care not separately coded")
add("Other consumption - Childcare services", ["12401"])
add("Other consumption - Education services", ["101"])
add("Other consumption - Religious and ritual articles", [], flag="NOMATCH", note="Not separately coded in COICOP at this resolution")
add("Other consumption - Preventive care services", [], flag="NOMATCH", note="Not separately coded; embedded within broader health codes")
add("Other consumption - Software", [], flag="NOMATCH", note="Not separately coded at this COICOP resolution; likely embedded in 0913")

# Leisure, sport and culture
add("Leisure, sport and culture - Package holidays", ["096"], flag="LOW-CONF",
    note="COICOP package-holiday spend understates footprint since it excludes the flight portion sold separately")
add("Leisure, sport and culture - Recreation and sport services", ["0941"])
add("Leisure, sport and culture - Cultural services", ["0942"])
add("Leisure, sport and culture - Games of chance, betting and lotteries", ["0943"])
add("Leisure, sport and culture - Games, toys and hobby articles", ["0931"])
add("Leisure, sport and culture - Garden products, plants and flowers", ["0933"])
add("Leisure, sport and culture - Pets and pet products", ["0934"])
add("Leisure, sport and culture - Services for pets", ["0935"])
add("Leisure, sport and culture - Equipment for sport, camping and outdoor recreation", ["0932"])
add("Leisure, sport and culture - Major durable goods for leisure", ["0921"])
add("Leisure, sport and culture - Musical instruments", ["0922"])
add("Leisure, sport and culture - Leisure goods: rental, maintenance and repair", ["0923"])
add("Leisure, sport and culture - Newspapers and periodicals", ["0952"])
add("Leisure, sport and culture - Books", ["0951"])
add("Leisure, sport and culture - Miscellaneous printed matter", ["0953"])
add("Leisure, sport and culture - Writing and drawing materials", ["0954"])
add("Leisure, sport and culture - Audiovisual media", ["0914"], flag="LOW-CONF",
    note="Legacy physical-media COICOP line; poor proxy for streaming-era audiovisual consumption")
add("Leisure, sport and culture - Photographic and optical equipment", ["0912"])

# Local transport
add("Local transport - Cars", ["0711"])
add("Local transport - Motorcycles", ["0712"])
add("Local transport - Bicycles", ["0713"])
add("Local transport - Fuels and lubricants", ["0722"])
add("Local transport - Spare parts and accessories", ["0721"])
add("Local transport - Maintenance and repairs", ["0723"])
add("Local transport - Driving lessons, driving licences and vehicle inspections", ["07241", "07242"])
add("Local transport - Road tolls", ["07243"])
add("Local transport - Parking services", ["07244"])
add("Local transport - Company-car benefit", ["07245"], flag="SHARED",
    note="07245 combined with Rental of personal transport equipment",
    peers=["Local transport - Rental of personal transport equipment"])
add("Local transport - Rental of personal transport equipment", ["07245"], flag="SHARED",
    note="07245 combined with Company-car benefit",
    peers=["Local transport - Company-car benefit"])
add("Local transport - Passenger transport by rail and tram", ["0731"], flag="LOW-CONF", note="Tram not separately coded; rail proxy used")
add("Local transport - Road transport services: car pools, buses and taxis", ["0732"])
add("Local transport - Passenger transport by sea", ["0734"])
add("Local transport - Public transport and other combined trips", ["0735"])
add("Local transport - Other passenger transport services", ["0736"], flag="SHARED",
    note="0736 combined with Other transport of goods", peers=["Local transport - Other transport of goods"])
add("Local transport - Other transport of goods", ["0736"], flag="SHARED",
    note="0736 combined with Other passenger transport services", peers=["Local transport - Other passenger transport services"])
add("Local transport - Postal and courier services", ["0811"])
add("Local transport - Direct emissions from vehicle use", [], flag="DIRECT",
    note="Physical tailpipe combustion (InsightOne/Transport Agency vehicle-km data), not COICOP expenditure. Use vehicle-km, not SEK.")

# Two categories exist in the full 111-category baseline (data/baseline_categories.csv)
# but are not among the 109 active categories in Table A2, since they carry no
# active reduction measure in Scenario 4. Included here for completeness of the
# crosswalk, flagged rather than left as unexplained gaps.
add("Other consumption - Laboratory services and imaging", ["0621"], flag="LOW-CONF",
    note="Inactive in Scenario 4 (baseline 0.085 kg/cap, negligible); proxy-matched to outpatient medical services")
add("Carbon capture", [], flag="NOT_APPLICABLE",
    note="Technology placeholder representing CCS deployment (baseline 0 by construction), not a household consumption category; no COICOP correspondence expected")

# Flights
add("Flights - Passenger flights", ["0733"], flag="PHYSICAL",
    note="COICOP 0733 drastically understates true air-travel activity (spend routes partly through package "
         "holidays and outbound spending abroad). Dawkins et al. (2024) source air-travel emissions from a "
         "separate physical dataset. We anchor flights to national per-capita air-travel distance instead "
         "of SEK: see FLIGHTS_KM_PER_CAPITA below.")

FLIGHTS_KM_PER_CAPITA = 5800  # Larsson et al. (2018), national average international travel distance
FLIGHTS_CATEGORY = "Flights - Passenger flights"


def build_crosswalk(scb, baseline_rows):
    """baseline_rows: list of dicts from data/baseline_categories.csv"""
    baseline = {r['category']: float(r['baseline_kgco2e_per_capita']) for r in baseline_rows}
    cid = {r['category']: r['category_id'] for r in baseline_rows}

    raw = {}
    for cat, info in M.items():
        codes = info['codes']
        sek_total = sum(scb[c][1] for c in codes) if codes else None
        raw[cat] = dict(info, sek_million=sek_total)

    out = []
    for cat, base_kg in baseline.items():
        info = raw.get(cat)
        if info is None:
            out.append(dict(category_id=cid[cat], category=cat, domain=cat.split(' - ')[0],
                             baseline_kgco2e_cap=base_kg, coicop_codes='', sek_per_capita_2021='',
                             implied_intensity_kgco2e_per_sek='', flag='MISSING_FROM_CROSSWALK',
                             note='Category not found in mapping - needs manual add'))
            continue

        flag, note = info['flag'], info['note']
        sek_pc, intensity = '', ''

        if cat == FLIGHTS_CATEGORY:
            sek_pc = FLIGHTS_KM_PER_CAPITA
            intensity = round(base_kg / FLIGHTS_KM_PER_CAPITA, 5)
        elif flag in ('OK', 'LOW-CONF') and info['sek_million'] is not None:
            sek_pc = round(info['sek_million'] * 1_000_000 / POP_2021, 1)
            if sek_pc > 0:
                intensity = round(base_kg / sek_pc, 5)
        elif flag == 'SHARED':
            members = sorted(set([cat] + info['peers']))
            total_emissions_share = sum(baseline.get(m, 0) for m in members)
            if total_emissions_share > 0 and info['sek_million'] is not None:
                this_share = base_kg / total_emissions_share
                split_sek_pc = info['sek_million'] * this_share * 1_000_000 / POP_2021
                sek_pc = round(split_sek_pc, 1)
                if split_sek_pc > 0:
                    intensity = round(base_kg / split_sek_pc, 5)
                note = note + f" | SEK/cap split proportionally to baseline emissions share across {len(members)} peers."

        out.append(dict(category_id=cid[cat], category=cat, domain=cat.split(' - ')[0],
                         baseline_kgco2e_cap=base_kg, coicop_codes=", ".join(info['codes']),
                         sek_per_capita_2021=sek_pc, implied_intensity_kgco2e_per_sek=intensity,
                         flag=flag, note=note))
    return out


def write_csv(path, rows, fieldnames=None):
    path = Path(path)
    if fieldnames is None:
        fieldnames = list(rows[0].keys()) if rows else []
    with path.open('w', newline='', encoding='utf-8') as f:
        w = csv.DictWriter(f, fieldnames=fieldnames)
        w.writeheader()
        for row in rows:
            w.writerow({k: row.get(k, '') for k in fieldnames})


def enrich_table_a2(crosswalk_rows):
    """Read the base Table A2 (written by extract_clean_dataset_from_workbook.py)
    and add the six activity-intensity columns, overwriting the file."""
    base_path = SUPP_DIR / 'table_a2_scenario4_parameters.csv'
    with base_path.open(newline='', encoding='utf-8') as f:
        base_rows = list(csv.DictReader(f))

    cw_by_cat = {r['category']: r for r in crosswalk_rows}
    DIRECT_LABELS = {
        "Housing - Direct emissions from heating": "Heating fuel combusted (direct)",
        "Housing - Direct household emissions": "Other household fuel combusted (direct)",
        "Local transport - Direct emissions from vehicle use": "Vehicle fuel combusted (direct)",
    }

    for row in base_rows:
        cat = row['Category']
        cw = cw_by_cat.get(cat)
        if cat == FLIGHTS_CATEGORY:
            row['Activity metric'] = "International air travel distance"
            row['Activity unit'] = "km/cap/year (national average)"
            row['Baseline activity M0,c'] = FLIGHTS_KM_PER_CAPITA
            row['Baseline intensity I0,c (kgCO2e/unit)'] = cw['implied_intensity_kgco2e_per_sek'] if cw else ''
            row['Activity data source'] = "Larsson et al. (2018); Kamb & Larsson (2019)"
            row['Confidence flag'] = "PHYSICAL - national literature average, not Malm\u00f6-specific"
        elif cat in DIRECT_LABELS:
            row['Activity metric'] = DIRECT_LABELS[cat]
            row['Activity unit'] = "physical fuel/vehicle-use volume (not separately quantified)"
            row['Baseline activity M0,c'] = ''
            row['Baseline intensity I0,c (kgCO2e/unit)'] = ''
            row['Activity data source'] = "Statistics Sweden direct-emissions series (physical, non-monetary)"
            row['Confidence flag'] = "DIRECT - already activity-linked in source data; SEK not applicable"
        elif cw is None or cw['flag'] == 'NOMATCH':
            row['Activity metric'] = ''
            row['Activity unit'] = ''
            row['Baseline activity M0,c'] = ''
            row['Baseline intensity I0,c (kgCO2e/unit)'] = ''
            row['Activity data source'] = "No COICOP counterpart at 4-digit resolution" if cw else ''
            base_kg = float(row['Baseline (kg CO\u2082e/cap)'])
            row['Confidence flag'] = f"NOMATCH - {base_kg:.3f} kg/cap; residual limitation" if cw else "MISSING - needs manual review"
        else:
            row['Activity metric'] = cw['coicop_codes']
            row['Activity unit'] = "SEK/cap/year (national average, 2021)"
            row['Baseline activity M0,c'] = cw['sek_per_capita_2021']
            row['Baseline intensity I0,c (kgCO2e/unit)'] = cw['implied_intensity_kgco2e_per_sek']
            row['Activity data source'] = "Statistics Sweden, Statistikdatabasen table 000000SG (2022)"
            flag = cw['flag']
            row['Confidence flag'] = f"{flag} - {cw['note']}" if flag in ('LOW-CONF', 'SHARED') else f"{flag} - direct COICOP match"

    fieldnames = list(base_rows[0].keys())
    write_csv(base_path, base_rows, fieldnames=fieldnames)
    return len(base_rows)


if __name__ == '__main__':
    if not SCB_WORKBOOK.exists():
        raise FileNotFoundError(
            f"Expected SCB workbook at {SCB_WORKBOOK}. Download table 000000SG from "
            "Statistikdatabasen (see module docstring) and place it in source/."
        )
    baseline_path = DATA_DIR / 'baseline_categories.csv'
    if not baseline_path.exists():
        raise FileNotFoundError("Run extract_clean_dataset_from_workbook.py first.")

    scb = read_scb_workbook(SCB_WORKBOOK)
    with baseline_path.open(newline='', encoding='utf-8') as f:
        baseline_rows = list(csv.DictReader(f))

    crosswalk = build_crosswalk(scb, baseline_rows)
    write_csv(DATA_DIR / 'coicop_activity_intensity_crosswalk.csv', crosswalk)
    print(f"Crosswalk: {len(crosswalk)} categories written to data/coicop_activity_intensity_crosswalk.csv")

    from collections import Counter
    print(Counter(r['flag'] for r in crosswalk))

    n = enrich_table_a2(crosswalk)
    print(f"Table A2 enriched with activity-intensity columns: {n} rows updated in supplementary/")
